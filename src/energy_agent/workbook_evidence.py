from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal

from energy_agent.evidence import HybridEvidenceIndex, OfficialChunk

FIGURE = re.compile(r"^Figure\s+(\d+)\s+(.+)$", re.IGNORECASE)


def _cell_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return f"{value:.12g}"
    return " ".join(str(value).split())


def _preview_rows(rows: list[tuple[object, ...]], max_rows: int, max_cells: int) -> tuple[str, int]:
    populated = [tuple(value for value in row if value not in (None, "")) for row in rows]
    populated = [row for row in populated if row]
    if len(populated) > max_rows:
        half = max_rows // 2
        populated = populated[:half] + populated[-(max_rows - half) :]
    rendered: list[str] = []
    used_cells = 0
    for row in populated:
        available = max_cells - used_cells
        if available <= 0:
            break
        selected = row[:available]
        rendered.append(" | ".join(_cell_text(value) for value in selected))
        used_cells += len(selected)
    return "\n".join(rendered), used_cells


@dataclass(frozen=True)
class FigureEvidence:
    chunk_id: str
    figure_id: str
    figure_number: int
    source_id: str
    title: str
    subtitle: str
    text: str
    url: str
    published_at: str
    retrieved_at: str
    sha256: str
    image_sha256: tuple[str, ...]
    image_count: int
    row_count: int
    column_count: int
    preview_cell_count: int
    evidence_type: str = "official_chart_workbook"

    def to_official_chunk(self) -> OfficialChunk:
        return OfficialChunk(
            chunk_id=self.chunk_id,
            source_id=self.source_id,
            title=f"{self.figure_id}: {self.title}",
            text=f"{self.subtitle}\n{self.text}",
            url=self.url,
            published_at=self.published_at,
            retrieved_at=self.retrieved_at,
            sha256=self.sha256,
            evidence_type=self.evidence_type,
        )

    def public_dict(self) -> dict[str, Any]:
        return asdict(self)


def extract_figure_evidence(
    payload: bytes,
    *,
    source_id: str,
    url: str,
    published_at: str,
    retrieved_at: str,
    max_preview_rows: int = 24,
    max_preview_cells: int = 400,
) -> list[FigureEvidence]:
    """Parse an official QED chart workbook into page-like figure evidence.

    AEMO QED workbooks pair each chart image with the underlying source cells.
    The index stores image hashes and a bounded tabular preview. It never treats
    OCR/VLM output as an observed market value.
    """

    if max_preview_rows < 2 or max_preview_cells < 10:
        raise ValueError("preview limits are too small")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised by dependency installation
        raise RuntimeError("openpyxl is required for QED workbook evidence") from exc

    workbook_sha256 = hashlib.sha256(payload).hexdigest()
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    if "Contents" not in workbook.sheetnames:
        raise ValueError("QED workbook is missing the Contents sheet")

    catalog: dict[int, tuple[str, str]] = {}
    for row in workbook["Contents"].iter_rows(values_only=True):
        for position, value in enumerate(row):
            if not isinstance(value, str):
                continue
            match = FIGURE.match(" ".join(value.split()))
            if not match:
                continue
            number = int(match.group(1))
            subtitle_value = row[position + 1] if position + 1 < len(row) else None
            subtitle = (
                " ".join(str(subtitle_value).split()) if subtitle_value not in (None, "") else ""
            )
            catalog[number] = (match.group(2), subtitle)
            break

    figures: list[FigureEvidence] = []
    for number in sorted(catalog):
        sheet_name = f"Figure {number}"
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Contents references missing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        preview, preview_cells = _preview_rows(rows[5:], max_preview_rows, max_preview_cells)
        image_hashes = tuple(
            hashlib.sha256(image._data()).hexdigest()
            for image in getattr(worksheet, "_images", [])
        )
        title, subtitle = catalog[number]
        figure_id = f"Figure {number}"
        figures.append(
            FigureEvidence(
                chunk_id=f"{source_id}-figure-{number:03d}",
                figure_id=figure_id,
                figure_number=number,
                source_id=source_id,
                title=title,
                subtitle=subtitle,
                text=preview,
                url=url,
                published_at=published_at,
                retrieved_at=retrieved_at,
                sha256=workbook_sha256,
                image_sha256=image_hashes,
                image_count=len(image_hashes),
                row_count=worksheet.max_row,
                column_count=worksheet.max_column,
                preview_cell_count=preview_cells,
            )
        )
    if not figures:
        raise ValueError("QED workbook contained no figure sheets")
    return figures


class FigureEvidenceIndex:
    """Auditable figure router over chart titles, captions and source cells."""

    backend = "qed_workbook_bm25+64d_lsa+rrf+numeric_rerank"

    def __init__(self, figures: list[FigureEvidence], dense_dimensions: int = 64) -> None:
        if not figures:
            raise ValueError("figures must not be empty")
        self.figures = figures
        self._by_chunk = {figure.chunk_id: figure for figure in figures}
        self._index = HybridEvidenceIndex(
            [figure.to_official_chunk() for figure in figures],
            dense_dimensions=dense_dimensions,
        )
        self.documents = self._index.documents

    def search(
        self,
        query: str,
        top_k: int = 5,
        mode: Literal["bm25", "dense", "rrf", "hybrid_rerank"] = "hybrid_rerank",
    ) -> list[dict[str, Any]]:
        hits = self._index.search(query, top_k=top_k, mode=mode)
        output: list[dict[str, Any]] = []
        for hit in hits:
            figure = self._by_chunk[str(hit["chunk_id"])]
            output.append(
                {
                    **hit,
                    "figure_id": figure.figure_id,
                    "figure_number": figure.figure_number,
                    "subtitle": figure.subtitle,
                    "image_sha256": list(figure.image_sha256),
                    "image_count": figure.image_count,
                    "row_count": figure.row_count,
                    "column_count": figure.column_count,
                    "preview_cell_count": figure.preview_cell_count,
                    "modality": "chart",
                    "asset_id": f"{figure.chunk_id}-image-001",
                    "asset_sha256": figure.image_sha256[0] if figure.image_sha256 else figure.sha256,
                }
            )
        return output

    def search_multimodal(
        self,
        query: str,
        top_k: int = 5,
        preferred_modality: Literal["auto", "text", "visual", "chart", "table"] = "auto",
    ) -> list[dict[str, Any]]:
        del preferred_modality
        return self.search(query, top_k)


def load_figure_evidence_records(path: Path) -> list[FigureEvidence]:
    """Load private, precompiled figure records without reopening source workbooks."""

    figures: list[FigureEvidence] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        row["image_sha256"] = tuple(row.get("image_sha256", []))
        figures.append(FigureEvidence(**row))
    return figures
