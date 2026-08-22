from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PillowImage

from energy_agent.market import fixture_store
from energy_agent.tools import ToolRegistry
from energy_agent.workbook_evidence import FigureEvidenceIndex, extract_figure_evidence


def fixture_workbook() -> bytes:
    workbook = Workbook()
    contents = workbook.active
    contents.title = "Contents"
    contents.append(["Quarterly Energy Dynamics"])
    contents.append(["Figure 1 Batteries shifted demand", "Demand by time of day"])
    contents.append(["Figure 2 Inter-regional transfers fell", "Quarterly transfers by link"])
    for number, header, values in [
        (1, ["hour", "battery MW"], [[17, 50.0], [18, 72.0]]),
        (2, ["link", "net flow MW"], [["VIC-SA", 320.5], ["NSW-QLD", -406.5]]),
    ]:
        sheet = workbook.create_sheet(f"Figure {number}")
        for row in [["Go to Contents"], ["QED"], ["Period"], [f"Figure {number}"], ["Subtitle"], header, *values]:
            sheet.append(row)
        bitmap = PillowImage.new("RGB", (2, 2), (number * 40, 20, 10))
        buffer = BytesIO()
        bitmap.save(buffer, format="PNG")
        buffer.seek(0)
        sheet.add_image(Image(buffer), "E2")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_extracts_typed_figure_data_and_image_provenance() -> None:
    figures = extract_figure_evidence(
        fixture_workbook(),
        source_id="qed",
        url="https://aemo.example/workbook.xlsx",
        published_at="2026-07-28",
        retrieved_at="2026-08-22T00:00:00+00:00",
    )
    assert [figure.figure_id for figure in figures] == ["Figure 1", "Figure 2"]
    assert figures[0].image_count == 1
    assert len(figures[0].image_sha256[0]) == 64
    assert "battery MW" in figures[0].text
    assert figures[1].preview_cell_count == 6


def test_figure_router_uses_caption_and_underlying_cells() -> None:
    figures = extract_figure_evidence(
        fixture_workbook(),
        source_id="qed",
        url="https://aemo.example/workbook.xlsx",
        published_at="2026-07-28",
        retrieved_at="2026-08-22T00:00:00+00:00",
    )
    index = FigureEvidenceIndex(figures, dense_dimensions=2)
    hit = index.search("VIC SA net flow 320.5 MW", top_k=1)[0]
    assert hit["figure_id"] == "Figure 2"
    assert hit["numeric_coverage"] > 0
    assert hit["modality"] == "chart"
    assert len(hit["asset_sha256"]) == 64


def test_typed_tool_routes_workbook_chart_without_exposing_a_file_path() -> None:
    figures = extract_figure_evidence(
        fixture_workbook(),
        source_id="qed",
        url="https://aemo.example/workbook.xlsx",
        published_at="2026-07-28",
        retrieved_at="2026-08-22T00:00:00+00:00",
    )
    registry = ToolRegistry(fixture_store(), FigureEvidenceIndex(figures, dense_dimensions=2))
    result = registry.execute(
        "search_official_evidence",
        {
            "query": "Show the VIC SA net flow chart",
            "retrieval_mode": "multimodal_fusion",
            "preferred_modality": "chart",
            "top_k": 1,
        },
    )
    assert result.evidence[0].modality == "chart"
    assert result.evidence[0].asset_id == "qed-figure-002-image-001"
    assert "path" not in result.evidence[0].model_dump()


def test_rejects_missing_contents_contract() -> None:
    workbook = Workbook()
    workbook.active.title = "Figure 1"
    output = BytesIO()
    workbook.save(output)
    try:
        extract_figure_evidence(
            output.getvalue(),
            source_id="qed",
            url="https://aemo.example/workbook.xlsx",
            published_at="2026-07-28",
            retrieved_at="2026-08-22T00:00:00+00:00",
        )
    except ValueError as exc:
        assert "Contents" in str(exc)
    else:
        raise AssertionError("missing Contents sheet must fail closed")
